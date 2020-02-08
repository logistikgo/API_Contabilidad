from django.shortcuts import render
from ApiViajes.models import PendientesEnviar, RelacionConceptoxProyecto, Ext_PendienteEnviar_Costo, Ext_PendienteEnviar_Precio
from openpyxl import load_workbook
from django import forms


class UploadFileForm(forms.Form):
	file = forms.FileField()

def import_data(request):
	if request.method == "POST":
		form = UploadFileForm(request.POST,
			request.FILES)
		if form.is_valid():
			wb = load_workbook(request.FILES['file'], data_only=True)
			ws = wb.active
			for i in range(2, ws.max_row+1): 
				Folio=ws.cell(row=i, column=1).value
				NombreCortoCliente =ws.cell(row=i, column=2).value
				NombreCortoProveedor=ws.cell(row=i, column=3).value
				FechaDescarga=ws.cell(row=i, column=4).value
				Moneda='MXN'
				CostoSubtotal=0
				CostoIVA=0
				CostoRetencion=0
				CostoTotal=0
				PrecioSubtotal=ws.cell(row=i, column=7).value
				PrecioIVA=ws.cell(row=i, column=8).value
				PrecioRetencion=ws.cell(row=i, column=9).value
				PrecioTotal=ws.cell(row=i, column=10).value
				PrecioServiciosIVA=ws.cell(row=i, column=11).value
				PrecioServiciosRetencion=ws.cell(row=i, column=12).value
				PrecioServiciosSubtotal=ws.cell(row=i, column=13).value
				PrecioServiciosTotal=ws.cell(row=i, column=14).value
				IsEvidenciaFisica = ws.cell(row=i, column=16).value 
				IsEvidenciaDigital = ws.cell(row=i, column=15).value 
				IsControlDesk = ws.cell(row=i, column=17).value 
				if IsEvidenciaFisica == 1:
					IsEvidenciaFisica = True
				else:
					IsEvidenciaFisica = False
				if IsEvidenciaDigital == 1:
					IsEvidenciaDigital = True
				else:
					IsEvidenciaDigital = False
				if IsControlDesk == 1:
					IsControlDesk = True
				else:
					IsControlDesk = False
				Estatus='FINALIZADO'
				IDProveedor=ws.cell(row=i, column=19).value
				IDCliente=ws.cell(row=i, column=20).value
				Proyecto = "BKG"
				TipoConcepto = "VIAJE"
				print(Folio)
				PendienteEnviar = PendientesEnviar(
					Folio=Folio,
					NombreCortoCliente=NombreCortoCliente,
					NombreCortoProveedor=NombreCortoProveedor,
					FechaDescarga=FechaDescarga,
					Moneda=Moneda,
					Status=Estatus,
					IsEvidenciaFisica=IsEvidenciaFisica,
					IsEvidenciaDigital=IsEvidenciaDigital,
					Proyecto=Proyecto,
					TipoConcepto=TipoConcepto,
					IsControlDesk=IsControlDesk
					)
				PendienteEnviar.save()
				Ext_Precio=Ext_PendienteEnviar_Precio(
					IDPendienteEnviar=PendienteEnviar,
					PrecioSubtotal=PrecioSubtotal,
					PrecioIVA=PrecioIVA,
					PrecioRetencion=PrecioRetencion,
					PrecioTotal=PrecioTotal,
					ServiciosIVA=PrecioServiciosIVA,
					ServiciosRetencion=PrecioServiciosRetencion,
					ServiciosSubtotal=PrecioServiciosSubtotal,
					ServiciosTotal=PrecioServiciosTotal
					)
				Ext_Precio.save()
				RelaConceptoxProyecto=RelacionConceptoxProyecto(
					IDPendienteEnviar = PendienteEnviar,
				    IDConcepto = 99999999,
				    IDCliente = IDCliente,
				    IDProveedor = IDProveedor
					)
				RelaConceptoxProyecto.save()
			return HttpResponse("OKKKK")
		else:
			return HttpResponseBadRequest()
	else:
		form = UploadFileForm()
	return render(
		request,
		'upload_form.html',
		{
		'form': form,
		'title': 'Import excel data into database example',
		'header': 'Please upload sample-data.xls:'
		})